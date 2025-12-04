import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
import globals as cfg
from datetime import date
import os

# Generate Excel Workbook containing unique values for each column in dataframe

today = cfg.get_today()
root_dir = cfg.get_project_root()
datasets = cfg.get_target_datasets()

today = date.today()
ts = today.strftime("%Y%m%d")

for dataset in datasets:
    # Set source and targets
    source_filename = dataset
    source_file = str(root_dir) + '/data/input/remapped/' + dataset
    target_path = str(root_dir) + '/data/output/unique-values/'
    datasetName = dataset.replace('.csv', '')
    if not os.path.isdir(target_path):
        os.mkdir(target_path)

    unique_counts_file = str(target_path) + '/' + datasetName + '-unique-values.xlsx'
    wb = Workbook()
    ws = wb.active
    # Create Contents Sheet
    ws0 = wb.create_sheet('Contents')
    worksheet = wb['Contents']
    worksheet['A1'] = "Unique Value Sets"
    worksheet['A2'] = "Each worksheet contains the unique values for a given column"
    worksheet['A4'] = "Created On: "
    worksheet['B4'] = today
    currentCell = worksheet['B4']
    currentCell.alignment = Alignment(horizontal='left')
    worksheet["A6"] = "Source Dataset: "
    worksheet['B7'] = datasetName
    worksheet['A9'] = "Columns"

    # Load Source Dataset
    df = pd.read_csv(source_file, sep=',', lineterminator='\n', encoding='utf-8', low_memory=False)
    print(df.columns)
    # Iterate over columns
    # for series_name, series in df.items():
    for col in df.columns:
        ws = wb.create_sheet(col)
        col_idx = df.columns.get_loc(col)
        # Get first 5 unique values then convert to list
        # unique_list = df[col].dropna().unique()[:5].tolist()
        # Print meta if more than one unique value
        unique_values = df[col].dropna().unique()

        for val in unique_values:
            print(val)
            ws.append([val])
    wb.save(unique_counts_file)
